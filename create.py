import calendar
from pathlib import Path
from typing import Optional

from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Border, Font, Alignment, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


FONT_BOLD = Font(bold=True, size="11", name="Calibri")
FONT_TITLE = Font(bold=True, size="18", name="Calibri")
ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
SIDE_BLACK = Side(border_style="thin", color="000000")
BORDER_BLACK = Border(top=SIDE_BLACK, bottom=SIDE_BLACK, left=SIDE_BLACK, right=SIDE_BLACK)
LEFT_BORDER_BLACK = Border(left=SIDE_BLACK)

BASE_PATH = Path("/home/lucifer/Downloads/CateringParis2025/template")
columns = [
    {"name": "Tour Code", "width": 16},
    {"name": "Tour Manager", "width": 16},
    {"name": "Service Date", "width": 14},
    {"name": "Service Type", "width": 14},
    {"name": "Adult", "width": 10},
    {"name": "Children", "width": 10},
    {"name": "Time", "width": 8},
    {"name": "DMC", "width": 16},
    {"name": "Total Pax", "width": 10},
    {"name": "Veg", "width": 10},
    {"name": "Non Veg", "width": 10},
    {"name": "Jain", "width": 10},
    {"name": "Delivery", "width": 10},
    {"name": "Remarks", "width": 10},
    {"name": "Price Adult", "width": 12},
    {"name": "Price Child", "width": 12},
]
restaurants = [
    ("Dawat", "Dawat"),
    ("WelcomeIndia", "Welcome India"),
    ("WaytoIndia", "Way to India"),
    ("Tara", "Tara")
]


def cell(ws, value, *,
         border: Optional[Border] = BORDER_BLACK,
         alignment: Optional[Alignment] = ALIGNMENT_CENTER,
         font: Optional[Font] = FONT_BOLD):
    _cell = WriteOnlyCell(ws, value)
    if border:
        _cell.border = border
    if alignment:
        _cell.alignment = alignment
    if font:
        _cell.font = font
    return _cell


def write_sheet(workbook, filename, restaurant):
    worksheet = workbook.create_sheet(title=restaurant[0])

    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[3].height = 25
    for idx, column in enumerate(columns):
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = column["width"]

    title_cell = cell(worksheet, restaurant[1].upper(), font=FONT_TITLE)
    date_cell = cell(worksheet, filename, font=FONT_TITLE)

    title_cell_right_border = WriteOnlyCell(worksheet, None)
    title_cell_right_border.border = LEFT_BORDER_BLACK

    address_cell_right_border = WriteOnlyCell(worksheet, None)
    address_cell_right_border.border = LEFT_BORDER_BLACK

    worksheet.append([title_cell, None, None, title_cell_right_border, None, date_cell, None, None, None, address_cell_right_border])
    worksheet.merged_cells.ranges.add("A1:C1")
    worksheet.merged_cells.ranges.add("F1:I1")

    worksheet.append([])
    header = [
        cell(worksheet, column["name"])
        for column in columns
    ]
    worksheet.append(header)


def write_file(month_name, day):
    filename = f"{day}-{month_name}"
    workbook = Workbook(write_only=True)
    for restaurant in restaurants:
        write_sheet(workbook, filename, restaurant)
    workbook.save(BASE_PATH / month_name / f"{filename}.xlsx")


def write_all_files():
    for month in range(1, 13):
        month_name = calendar.month_name[month]
        month_path = BASE_PATH / month_name
        month_path.mkdir(parents=True, exist_ok=True)

        _, last_day = calendar.monthrange(2025, month)
        for day in range(1, last_day + 1):
            write_file(month_name, day)


if __name__ == "__main__":
    write_all_files()
