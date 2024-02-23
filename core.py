import os
import traceback
from datetime import datetime, date
from enum import Enum

from Levenshtein import distance
from dateutil.relativedelta import relativedelta
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from pandas import DataFrame, concat, Timestamp, options

from openpyxl import load_workbook

options.mode.copy_on_write = True

MAX_ROWS = 50
MAX_COLS = 25


class Restaurant(Enum):
    DAWAT = "Dawat"
    WELCOME_INDIA = "WelcomeIndia"


RESTAURANTS = [Restaurant.DAWAT, Restaurant.WELCOME_INDIA]


CANONICAL_DMCS = {
    "Neemholidays": "Neem Holidays",
    "Star Our": "Star Tour",
    "Star": "Star Tour",
    "Ezxa": "Exza",
    "Tc": "TC",
    "Gtt": "GTT",
    "Chr": "CHR",
    "Gb Dmc": "GB DMC",
    "Gb Dmc Ltd.": "GB DMC",
    "Youngedsplorer": "Young Edsplorer",
    "Truvaiglobal": "Truvai",
    "Truvai Dmc": "Truvai",
    "Truvai Global": "Truvai",
    "Europe Incomign": "Europe Incoming",
    "Europeincoming": "Europe Incoming",
    "Afc Holidyays": "AFC Holidays",
    "Afc Holidays": "AFC Holidays",
    "G2 Travel": "G2 Travels",
    "G2Travel": "G2 Travels",
    "G2-Travel": "G2 Travels",
    "Holiday Carnival": "Holidays Carnival",
    "Europe Goodlife": "Europe Good Life",
    "Europegoodlife": "Europe Good Life",
    "Gateways Group Of Dmcs": "Gateways Group Of Dmc'S",
    "Gateways Dmc": "Gateways Group Of Dmc'S",
    "Gateways Group Of Dmc’S": "Gateways Group Of Dmc'S",
    "Deewan Holidays": "Dewan Holidays",
    "Dewan Travels": "Dewan Holidays",
    "Switru": "Switrus",
    "European Gatewyas": "European Gateways",
    "Lamour Voyage": "Lamour Voyages",
    "Lamondialetour": "La Mondiale",
    "Lamondiale Tour": "La Mondiale",
    "Lamondiale Tours": "La Mondiale",
    "Mahadevan Group": "Mahadevan",
    "Whats App": "WhatsApp",
    "Whatassp": "WhatsApp",
    "Whatsapp": "WhatsApp"
}
IGNORE_DMC_DUPES = [("TC", "Tcf"), ("Magi Holidays", "Mango Holidays"), ("GTT", "TC")]


def get_directories(from_date: date, to_date: date):
    months_done = set()
    directories = []

    delta = relativedelta(months=1, day=1)
    current_date = from_date
    while current_date <= to_date:
        if current_date.month not in months_done:
            directories.append(current_date.strftime("%B"))
            directories.append(current_date.strftime("%b"))
        months_done.add(current_date.month)
        current_date += delta

    return directories


def list_files(updater, from_date, to_date, base_dir):
    directories = get_directories(from_date, to_date)
    updater(f"Searching for directories: {', '.join(directories)}")

    files = []
    for directory in get_directories(from_date, to_date):
        dir_path = os.path.join(base_dir, directory)
        if not os.path.exists(dir_path):
            continue

        for file in os.listdir(dir_path):
            if not (file.endswith(".xlsx") or file.endswith("xls")):
                continue

            parts = file.replace("-", " ").split()
            try:
                day = int(parts[0])
            except ValueError:
                continue

            files.append(f"{base_dir}/{directory}/{file}")

    return files


def get_sheet_contents(sheet):
    header = None
    data = []

    itr = sheet.iter_rows(max_col=MAX_COLS, max_row=MAX_ROWS, values_only=True)
    while (row := next(itr, None)) is not None:
        if row[0] is not None and row[0].lower() == "tour code":
            header = list(row)
            while header and header[-1] is None:
                header.pop()
            break

    if not header:
        return None, None

    header = [column.title() for column in header]

    n_columns = len(header)

    while (row := next(itr, None)) is not None:
        values = row[:n_columns]
        if not all(value is None for value in values):
            data.append(list(values))

    return header, data or None


def get_file_contents(file) -> DataFrame | None:
    dfs = []

    workbook = load_workbook(file, read_only=True)
    for restaurant in RESTAURANTS:
        sheet = workbook[restaurant.value]
        header, data = get_sheet_contents(sheet)
        if data is None:
            continue
        df = DataFrame(data, columns=header)
        df = df.dropna(axis=1, how="all")
        df["Restaurant"] = restaurant.value
        dfs.append(df)
    workbook.close()

    if len(dfs) == 0:
        return None

    return concat(dfs, ignore_index=True)


def get_all_data(updater, from_date, to_date, base_dir):
    files = list_files(updater, from_date, to_date, base_dir)
    updater("Found {} files".format(len(files)))

    dfs = []
    for file in files:
        df = get_file_contents(file)
        if df is None:
            updater("No data found for " + file)
            continue
        dfs.append(df)

    if len(dfs) == 0:
        updater("No data found for any file")
        return None

    combined_df = concat(dfs, ignore_index=True)
    return combined_df


def convert_to_date(value):
    if isinstance(value, Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def find_possible_dupes(updater, dmcs):
    for dmc1 in dmcs:
        dupes = []
        for dmc2 in dmcs:
            if dmc1 == dmc2:
                continue
            if distance(dmc1, dmc2) <= 2\
                    and (dmc1, dmc2) not in IGNORE_DMC_DUPES\
                    and (dmc2, dmc1) not in IGNORE_DMC_DUPES:
                dupes.append(dmc2)
        if dupes:
            updater(f"Possible duplicates for '{dmc1}': {dupes}")


def process_data(updater, from_date, to_date, df: DataFrame) -> tuple[DataFrame, DataFrame]:
    df["Service Date Cleaned"] = df["Service Date"].apply(convert_to_date)

    df = df.loc[(df["Service Date Cleaned"] >= from_date) & (df["Service Date Cleaned"] <= to_date)]

    df["Adult"] = df["Adult"].fillna(0)
    df["Children"] = df["Children"].fillna(0)

    df["Dmc Canonical"] = df["Dmc"]
    df["Dmc Canonical"] = df["Dmc Canonical"].str.title().str.strip().str.split().str.join(" ")
    df["Dmc Canonical"] = df["Dmc Canonical"].apply(lambda dmc: CANONICAL_DMCS.get(dmc, dmc))

    # TODO: Create a list of valid DMCs, write invalid ones to a separate file
    unique_dmcs = df["Dmc Canonical"].unique().tolist()
    if None in unique_dmcs:
        unique_dmcs.remove(None)
    find_possible_dupes(updater, unique_dmcs)

    if "Price Adult" in df.columns:
        df["Price Adult"] = df["Price Adult"].fillna(13)

    if "Price Child" in df.columns:
        df["Price Child"] = df["Price Child"].fillna(8)

    if "Remarks" in df.columns:
        cancelled_mask_1 = df["Remarks"].fillna("").str.lower().str.startswith("cancel")
    else:
        cancelled_mask_1 = True

    if "Delivery" in df.columns:
        cancelled_mask_2 = df["Delivery"].fillna("").str.lower().str.startswith("cancel")
    else:
        cancelled_mask_2 = True

    cancelled_mask = cancelled_mask_1 | cancelled_mask_2

    serviced_df = df[~cancelled_mask]
    cancelled_df = df[cancelled_mask]

    return serviced_df, cancelled_df


def write_invoice(updater, workbook, save_path, group: DataFrame):
    sorted_group = group.sort_values("Service Date Cleaned")

    offset = 1

    ws = workbook.create_sheet()

    columns = ["Tour Code", "Tour Manager", "Service Date", "Service Type", "Adult", "Children", "Price Adult", "Price Child"]
    write_df = sorted_group[columns]
    write_df["Service Date To Print"] = sorted_group.apply(lambda r: r["Service Date Cleaned"] or r["Service Date"], axis=1)

    # itertuples cannot handle space in column name properly
    write_df = write_df.rename(lambda x: x.replace(" ", "_"), axis="columns")

    columns.append("Total")
    ws.append(columns)
    offset += 1

    for idx, row in enumerate(write_df.itertuples(index=False)):
        n_row = idx + offset
        service_date_cell = WriteOnlyCell(ws, row.Service_Date_To_Print)
        service_date_cell.number_format = "dd mmm"
        total_formula = f"=E{n_row} * G{n_row} + F{n_row} * H{n_row}"
        ws.append([
            row.Tour_Code,
            row.Tour_Manager,
            service_date_cell,
            row.Service_Type,
            row.Adult,
            row.Children,
            row.Price_Adult,
            row.Price_Child,
            total_formula
        ])

    ws.append([])

    grand_total_cell = WriteOnlyCell(ws, f"=SUM(I{offset}:I{offset + write_df.shape[0]})")
    ws.append([None, None, None, "Grant Total", None, None, None, None, grand_total_cell])

    workbook.save(save_path)

    updater(f"Saved invoice to {save_path}")


def write_all_invoices(updater, base_dir, restaurant, df: DataFrame):
    dir_path = os.path.join(base_dir, restaurant.value, "AutoInvoices")
    os.makedirs(dir_path, exist_ok=True)

    for name, group in df.groupby("Dmc Canonical"):
        try:
            workbook = Workbook(write_only=True)
            save_path = os.path.join(dir_path, str(name) + ".xlsx")
            write_invoice(updater, workbook, save_path, group)
        except Exception:
            workbook.close()
            updater("Unable to write invoice for {name}: {e}".format(name=name, e=traceback.format_exc()))


def write_cancelled_df(updater, base_dir, restaurant, df: DataFrame):
    dir_path = os.path.join(base_dir, restaurant.value)
    os.makedirs(dir_path, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    save_path = os.path.join(dir_path, "Cancelled.xlsx")
    wb.save(save_path)
    updater(f"Saved cancelled tours to {save_path}")
