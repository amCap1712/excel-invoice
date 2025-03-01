import os
import traceback
from datetime import date
from typing import Optional

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl.cell import WriteOnlyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from pandas import concat, DataFrame

from .core import RESTAURANTS

MAX_ROWS = 50
MAX_COLS = 25
FONT_BOLD = Font(bold=True)
ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
SIDE_BLACK = Side(border_style="thin", color="000000")
BORDER_BLACK = Border(top=SIDE_BLACK, bottom=SIDE_BLACK, left=SIDE_BLACK, right=SIDE_BLACK)


def get_directories(from_date: date, to_date: date):
    months_done = set()
    directories = []

    delta = relativedelta(months=1, day=1)
    current_date = from_date
    while current_date <= to_date:
        month = current_date.strftime("%B")
        if month not in months_done:
            directories.append(month)
            months_done.add(month)

        month = current_date.strftime("%b")
        if month not in months_done:
            directories.append(month)
            months_done.add(month)
        current_date += delta

    return directories


def list_files(updater, from_date, to_date, base_dir):
    directories = get_directories(from_date, to_date)
    updater(f"Searching for directories: {', '.join(directories)}")

    files = []
    for directory in get_directories(from_date, to_date):
        dir_path = os.path.join(base_dir, directory)
        if not os.path.exists(dir_path):
            continue

        for file in os.listdir(dir_path):
            if not (file.endswith(".xlsx") or file.endswith("xls")):
                continue

            parts = file.replace("-", " ").split()
            try:
                day = int(parts[0])
            except ValueError:
                continue

            files.append(f"{base_dir}/{directory}/{file}")

    return files


def read_sheet(sheet):
    header = None
    data = []

    itr = sheet.iter_rows(max_col=MAX_COLS, max_row=MAX_ROWS, values_only=True)
    while (row := next(itr, None)) is not None:
        if row[0] is not None and row[0].lower() == "tour code":
            header = list(row)
            while header and header[-1] is None:
                header.pop()
            break

    if not header:
        return None, None

    header = [column.title() for column in header]

    n_columns = len(header)

    while (row := next(itr, None)) is not None:
        values = row[:n_columns]
        if not all(value is None for value in values):
            data.append(list(values))

    return header, data or None


def read_file(file) -> tuple[DataFrame | None, list[tuple], list[tuple]]:
    filename = os.path.basename(file)
    not_found = []
    typos = []
    dfs = []

    workbook = load_workbook(file, read_only=True)
    for restaurant in RESTAURANTS:
        sheet = None
        if restaurant.name in workbook:
            sheet = workbook[restaurant.name]
        else:
            for sheetname in workbook.sheetnames:
                if (restaurant.name == "Dawat" and sheetname.lower().startswith("d")) or \
                        (restaurant.name == "WelcomeIndia" and sheetname.lower().startswith("wel")) or \
                        (restaurant.name == "WaytoIndia" and sheetname.lower().startswith("way")) or \
                        (restaurant.name == "Tara" and sheetname.lower().startswith("t")):
                    sheet = workbook[sheetname]
                    typos.append((filename, restaurant.name, sheetname))
                    break

        if sheet is None:
            not_found.append((filename, restaurant.name))
            continue

        header, data = read_sheet(sheet)
        if data is None:
            continue
        df = DataFrame(data, columns=header)
        df = df.dropna(axis=1, how="all")
        df["Restaurant"] = restaurant.name

        df["File Name"] = filename
        df.insert(0, "File Name", df.pop("File Name"))

        dfs.append(df)
    workbook.close()

    if len(dfs) == 0:
        return None, not_found, typos

    return concat(dfs, ignore_index=True), not_found, typos


def read_all_files(updater, from_date, to_date, base_dir):
    files = list_files(updater, from_date, to_date, base_dir)
    updater(f"Found {len(files)} files")

    dfs = []
    not_found = []
    typos = []
    for file in files:
        df, _not_found, _typos = read_file(file)
        not_found.extend(_not_found)
        typos.extend(_typos)
        if df is None:
            updater("No data found for " + file)
            continue
        dfs.append(df)

    not_found_df = DataFrame(not_found, columns=["File Name", "Restaurant"]).drop_duplicates(ignore_index=True)
    typos_df = DataFrame(typos, columns=["File Name", "Restaurant", "Sheet Name"]).drop_duplicates(ignore_index=True)

    if len(dfs) == 0:
        updater("No data found for any file")
        return None, not_found_df, typos_df

    combined_df = concat(dfs, ignore_index=True).drop_duplicates(ignore_index=True)
    return combined_df, not_found_df, typos_df


def read_rates_file(base_dir):
    return pd.read_excel(os.path.join(base_dir, "Rates.xlsx"))


def cell(ws, value, *,
         border: Optional[Border] = BORDER_BLACK,
         alignment: Optional[Alignment] = ALIGNMENT_CENTER,
         font: Optional[Font] = None):
    _cell = WriteOnlyCell(ws, value)
    if border:
        _cell.border = border
    if alignment:
        _cell.alignment = alignment
    if font:
        _cell.font = font
    return _cell


def ecell(ws):
    return cell(ws, None, border=BORDER_BLACK, alignment=None, font=None)


def write_invoice(updater, workbook, save_path, address, dmc, group: DataFrame):
    sorted_group = group.sort_values("Service Date Cleaned")

    offset = 1
    ws = workbook.create_sheet()

    # column dimensions need to be written before any cell is written
    for idx in range(1, 10):
        if idx == 1:
            width = 20
        elif idx <= 3:
            width = 16
        else:
            width = 13
        ws.column_dimensions[get_column_letter(idx)].width = width

    address_cell = cell(ws, address, font=FONT_BOLD, alignment=Alignment(wrapText=True, vertical="top"),
                        border=BORDER_BLACK)
    dmc_cell = cell(ws, dmc, font=FONT_BOLD, alignment=Alignment(horizontal="center", vertical="top"),
                    border=BORDER_BLACK)

    ws.append([address_cell, None, None, None, dmc_cell, ecell(ws), ecell(ws), ecell(ws)])
    ws.append([ecell(ws), None, None, None, ecell(ws), ecell(ws), ecell(ws), ecell(ws)])
    ws.append([ecell(ws), None, None, None, ecell(ws), ecell(ws), ecell(ws), ecell(ws)])
    ws.merged_cells.ranges.add("A1:B3")
    ws.merged_cells.ranges.add("E1:H3")

    ws.append([])

    invoice_number_cell = cell(ws, "Invoice No.", font=FONT_BOLD, alignment=None, border=None)
    date_cell = cell(ws, "Date", font=FONT_BOLD, alignment=None, border=None)
    ws.append([invoice_number_cell, None, None, None, None, date_cell])
    ws.merged_cells.ranges.add("A5:B5")
    ws.merged_cells.ranges.add("G5:H5")

    ws.append([])
    offset += 6

    columns = ["Tour Code", "Service Date", "Service Type",
               "Adult", "Children", "Price Adult", "Price Child"]

    write_df = sorted_group[columns]
    # itertuples cannot handle space in column name properly
    write_df = write_df.rename(lambda x: x.replace(" ", "_"), axis="columns")

    columns.append("Total")
    column_cells = []
    for column in columns:
        column_cell = cell(ws, column, font=FONT_BOLD)
        column_cells.append(column_cell)
    ws.append(column_cells)
    offset += 1

    for idx, row in enumerate(write_df.itertuples(index=False)):
        row_cells = []
        for row_idx, value in enumerate(row):
            if row_idx == 1:
                row_cell = cell(ws, row.Service_Date)
                row_cell.number_format = "dd mmm"
            else:
                row_cell = cell(ws, value, font=None)
            row_cells.append(row_cell)

        n_row = idx + offset
        total_formula_cell = cell(ws, f"=D{n_row} * F{n_row} + E{n_row} * G{n_row}")
        row_cells.append(total_formula_cell)

        ws.append(row_cells)

    ws.append([ecell(ws), ecell(ws), ecell(ws), ecell(ws),
               ecell(ws), ecell(ws), ecell(ws), ecell(ws)])

    grand_total_label_cell = cell(ws, "Grand Total", font=FONT_BOLD)
    grand_total_cell = cell(ws, f"=SUM(H{offset}:H{offset + write_df.shape[0]})", font=FONT_BOLD)
    ws.append([ecell(ws), ecell(ws), ecell(ws), ecell(ws),
               ecell(ws), ecell(ws), grand_total_label_cell, grand_total_cell])

    workbook.save(save_path)
    updater(f"Saved invoice to {save_path}")


def write_all_invoices(updater, base_dir, address, df: DataFrame):
    for name, group in df.groupby("Dmc Canonical"):
        try:
            workbook = Workbook(write_only=True)
            save_path = os.path.join(base_dir, str(name) + ".xlsx")
            write_invoice(updater, workbook, save_path, address, name, group)
        except Exception:
            workbook.close()
            updater("Unable to write invoice for {name}: {e}".format(name=name, e=traceback.format_exc()))


def write_auxiliary_df(updater, base_dir, name, df: DataFrame):
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for idx in range(1, len(df.columns.tolist()) + 1):
        width = 30 if idx == 1 else 18
        ws.column_dimensions[get_column_letter(idx)].width = width
    save_path = os.path.join(base_dir, f"{name}.xlsx")
    wb.save(save_path)
    updater(f"Saved {name} tours to {save_path}")
