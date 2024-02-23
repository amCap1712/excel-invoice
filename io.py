import os
import traceback
from datetime import date

from dateutil.relativedelta import relativedelta
from openpyxl.cell import WriteOnlyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from pandas import concat, DataFrame

from core import RESTAURANTS


MAX_ROWS = 50
MAX_COLS = 25


def get_directories(from_date: date, to_date: date):
    months_done = set()
    directories = []

    delta = relativedelta(months=1, day=1)
    current_date = from_date
    while current_date <= to_date:
        if current_date.month not in months_done:
            directories.append(current_date.strftime("%B"))
            directories.append(current_date.strftime("%b"))
        months_done.add(current_date.month)
        current_date += delta

    return directories


def list_files(updater, from_date, to_date, base_dir):
    directories = get_directories(from_date, to_date)
    updater(f"Searching for directories: {', '.join(directories)}")

    files = []
    for directory in get_directories(from_date, to_date):
        dir_path = os.path.join(base_dir, directory)
        if not os.path.exists(dir_path):
            continue

        for file in os.listdir(dir_path):
            if not (file.endswith(".xlsx") or file.endswith("xls")):
                continue

            parts = file.replace("-", " ").split()
            try:
                day = int(parts[0])
            except ValueError:
                continue

            files.append(f"{base_dir}/{directory}/{file}")

    return files


def get_sheet_contents(sheet):
    header = None
    data = []

    itr = sheet.iter_rows(max_col=MAX_COLS, max_row=MAX_ROWS, values_only=True)
    while (row := next(itr, None)) is not None:
        if row[0] is not None and row[0].lower() == "tour code":
            header = list(row)
            while header and header[-1] is None:
                header.pop()
            break

    if not header:
        return None, None

    header = [column.title() for column in header]

    n_columns = len(header)

    while (row := next(itr, None)) is not None:
        values = row[:n_columns]
        if not all(value is None for value in values):
            data.append(list(values))

    return header, data or None


def get_file_contents(file) -> DataFrame | None:
    dfs = []

    workbook = load_workbook(file, read_only=True)
    for restaurant in RESTAURANTS:
        sheet = workbook[restaurant.value]
        header, data = get_sheet_contents(sheet)
        if data is None:
            continue
        df = DataFrame(data, columns=header)
        df = df.dropna(axis=1, how="all")
        df["Restaurant"] = restaurant.value
        dfs.append(df)
    workbook.close()

    if len(dfs) == 0:
        return None

    return concat(dfs, ignore_index=True)


def read_all_data(updater, from_date, to_date, base_dir):
    files = list_files(updater, from_date, to_date, base_dir)
    updater("Found {} files".format(len(files)))

    dfs = []
    for file in files:
        df = get_file_contents(file)
        if df is None:
            updater("No data found for " + file)
            continue
        dfs.append(df)

    if len(dfs) == 0:
        updater("No data found for any file")
        return None

    combined_df = concat(dfs, ignore_index=True)
    return combined_df


def write_invoice(updater, workbook, save_path, group: DataFrame):
    # TODO: write invoice header
    sorted_group = group.sort_values("Service Date Cleaned")

    offset = 1

    ws = workbook.create_sheet()

    columns = ["Tour Code", "Tour Manager", "Service Date", "Service Type", "Adult", "Children", "Price Adult", "Price Child"]
    write_df = sorted_group[columns]
    write_df["Service Date To Print"] = sorted_group.apply(lambda r: r["Service Date Cleaned"] or r["Service Date"], axis=1)

    # itertuples cannot handle space in column name properly
    write_df = write_df.rename(lambda x: x.replace(" ", "_"), axis="columns")

    columns.append("Total")
    ws.append(columns)
    offset += 1

    for idx, row in enumerate(write_df.itertuples(index=False)):
        n_row = idx + offset
        service_date_cell = WriteOnlyCell(ws, row.Service_Date_To_Print)
        service_date_cell.number_format = "dd mmm"
        total_formula = f"=E{n_row} * G{n_row} + F{n_row} * H{n_row}"
        ws.append([
            row.Tour_Code,
            row.Tour_Manager,
            service_date_cell,
            row.Service_Type,
            row.Adult,
            row.Children,
            row.Price_Adult,
            row.Price_Child,
            total_formula
        ])

    ws.append([])

    grand_total_cell = WriteOnlyCell(ws, f"=SUM(I{offset}:I{offset + write_df.shape[0]})")
    ws.append([None, None, None, "Grant Total", None, None, None, None, grand_total_cell])

    workbook.save(save_path)

    updater(f"Saved invoice to {save_path}")


def write_all_invoices(updater, base_dir, df: DataFrame):
    for name, group in df.groupby("Dmc Canonical"):
        try:
            workbook = Workbook(write_only=True)
            save_path = os.path.join(base_dir, str(name) + ".xlsx")
            write_invoice(updater, workbook, save_path, group)
        except Exception:
            workbook.close()
            updater("Unable to write invoice for {name}: {e}".format(name=name, e=traceback.format_exc()))


def write_cancelled_df(updater, base_dir, df: DataFrame):
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    save_path = os.path.join(base_dir, "Cancelled.xlsx")
    wb.save(save_path)
    updater(f"Saved cancelled tours to {save_path}")
